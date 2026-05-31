"""
Export database to ambassadors_base.xlsx format.

Sheet structure: All + one sheet per manager.
Columns: Channel | Sp ID | USDT-TRC20 | PayPal | Payment Method |
         Manager | Language | Platform | Content | Old Addresses
"""
from __future__ import annotations
import io
import os
from datetime import datetime

from telegram import Update
from telegram.ext import ContextTypes, CommandHandler, CallbackQueryHandler

from database.queries import (
    get_user, get_all_bloggers, get_active_methods,
    get_all_method_history, METHOD_LABELS,
)
from handlers.common import admin_only, get_lang
from config import ADMIN_ID

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# --------------------------------------------------------------------------- #
# Styling — matches ambassadors_base format
# --------------------------------------------------------------------------- #
HDR_FILL  = PatternFill("solid", fgColor="1F3864") if HAS_OPENPYXL else None
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11) if HAS_OPENPYXL else None
HDR_ALIGN = Alignment(horizontal="center", vertical="center") if HAS_OPENPYXL else None
ALT_FILL  = PatternFill("solid", fgColor="EEF2FF") if HAS_OPENPYXL else None
NORM_FILL = PatternFill("solid", fgColor="FFFFFF") if HAS_OPENPYXL else None
CELL_ALIG = Alignment(vertical="center", wrap_text=False) if HAS_OPENPYXL else None
OLD_ALIG  = Alignment(vertical="center", wrap_text=True) if HAS_OPENPYXL else None

COLUMNS = [
    ("Channel",        22),
    ("Sp ID",          26),
    ("USDT-TRC20",     36),
    ("PayPal",         30),
    ("Payment Method", 16),
    ("Manager",        18),
    ("Language",       14),
    ("Platform",       18),
    ("Content",        20),
    ("Old Addresses",  40),
]


# --------------------------------------------------------------------------- #
# Build data
# --------------------------------------------------------------------------- #
async def _build_export_data() -> dict[str, list[dict]]:
    """
    Returns {"All": [...], "John": [...], ...}
    Each item is a row dict with keys matching COLUMNS.
    """
    bloggers   = await get_all_bloggers()
    all_history = await get_all_method_history()  # {blogger_id: [{type, address}]}

    rows_by_manager: dict[str, list[dict]] = {}
    all_rows: list[dict] = []

    for b in bloggers:
        methods  = await get_active_methods(b["id"])
        history  = all_history.get(b["id"], [])

        # Current methods by type
        mmap = {m["type"]: m["address"] for m in methods}
        site  = mmap.get("site", "")
        usdt  = mmap.get("usdt-trc20", "")
        paypal = mmap.get("paypal", "")

        # Primary method label
        primary = next(
            (m for m in methods if m.get("is_primary")),
            methods[0] if methods else None,
        )
        pay_method = METHOD_LABELS.get(primary["type"], primary["type"]) if primary else ""

        # Old addresses from history
        old_parts = []
        for h in history:
            label = METHOD_LABELS.get(h["type"], h["type"])
            old_parts.append(f"{label}: {h['address']}")
        old_addresses = "\n".join(old_parts)

        # Manager name from users table via manager_id
        manager_name = b.get("manager_username") or b.get("manager_name") or ""

        row = {
            "Channel":        b["name"],
            "Sp ID":          site,
            "USDT-TRC20":     usdt,
            "PayPal":         paypal,
            "Payment Method": pay_method,
            "Manager":        manager_name,
            "Language":       "",
            "Platform":       "",
            "Content":        "",
            "Old Addresses":  old_addresses,
        }
        all_rows.append(row)
        rows_by_manager.setdefault(manager_name, []).append(row)

    result = {"All": sorted(all_rows, key=lambda r: r["Channel"].lower())}
    for mgr, rows in sorted(rows_by_manager.items()):
        if mgr:
            result[mgr] = sorted(rows, key=lambda r: r["Channel"].lower())

    return result


# --------------------------------------------------------------------------- #
# Write xlsx
# --------------------------------------------------------------------------- #
def _write_header(ws):
    for ci, (name, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.fill      = HDR_FILL
        c.font      = HDR_FONT
        c.alignment = HDR_ALIGN
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"


def _write_row(ws, row_num: int, row: dict):
    fill = ALT_FILL if row_num % 2 == 0 else NORM_FILL
    for ci, (col_name, _) in enumerate(COLUMNS, 1):
        val = row.get(col_name, "")
        c = ws.cell(row=row_num, column=ci, value=val)
        c.fill      = fill
        c.alignment = OLD_ALIG if col_name == "Old Addresses" else CELL_ALIG


def _build_xlsx(data: dict[str, list[dict]]) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, rows in data.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        _write_header(ws)
        for i, row in enumerate(rows, 2):
            _write_row(ws, i, row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Handler
# --------------------------------------------------------------------------- #
async def _do_export(update, context, lang: str):
    if not HAS_OPENPYXL:
        msg = (
            "Модуль openpyxl не установлен. Выполни на сервере:\npip install openpyxl"
            if lang == "ru" else
            "Module openpyxl is not installed. Run on server:\npip install openpyxl"
        )
        await update.effective_chat.send_message(msg)
        return

    await update.effective_chat.send_message(
        "Генерирую таблицу..." if lang == "ru" else "Generating spreadsheet..."
    )
    data  = await _build_export_data()
    xlsx  = _build_xlsx(data)
    ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"ambassadors_base_{ts}.xlsx"
    total = sum(len(v) for k, v in data.items() if k != "All")

    await update.effective_chat.send_document(
        document=io.BytesIO(xlsx),
        filename=fname,
        read_timeout=120,
        write_timeout=120,
        caption=(
            f"База амбассадоров · {len(data['All'])} блогеров · {len(data)-1} менеджеров"
            if lang == "ru" else
            f"Ambassadors base · {len(data['All'])} bloggers · {len(data)-1} managers"
        ),
    )


@admin_only
async def cmd_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = await get_user(update.effective_user.id)
    lang = get_lang(user)
    await _do_export(update, context, lang)


async def cb_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user = await get_user(update.effective_user.id)
    if not user or user["telegram_id"] != ADMIN_ID:
        await query.answer("Access denied.", show_alert=True)
        return
    lang = get_lang(user)
    await _do_export(update, context, lang)


# --------------------------------------------------------------------------- #
# Registration
# --------------------------------------------------------------------------- #
def register_export_handlers(app):
    app.add_handler(CommandHandler("export", cmd_export))
    app.add_handler(CallbackQueryHandler(cb_export, pattern=r"^adm:export$"))